import traceback
import pandas as pd
import xlrd
from datetime import datetime
from rdkit import Chem
from rdkit.Chem import Draw
from tqdm import tqdm
from utils import ymlReader
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import os
import logging

# 初始化日志，输出到控制台
log = logging.getLogger("DataMerger")
console_handler = logging.StreamHandler()
console_handler.setLevel('INFO')
fmt = u'%(asctime)s - %(funcName)s(): %(lineno)s [%(levelname)s]: %(message)s'
formatter = logging.Formatter(fmt)
console_handler.setFormatter(formatter)

log.setLevel('INFO')
log.addHandler(console_handler)


def xls2xlsx(xls_path: str) -> str:
    """
    将xls文件转换成xlsx
    :param xls_path: xls文件路径
    :return: 保存完成的xlsx路径
    """
    # 使用xlrd打开xls工资表
    book = xlrd.open_workbook(xls_path)
    index = 0
    nrows, ncols = 0, 0
    sheet = book.sheet_by_index(0)
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # 使用openpyxl准备一个xlsx工作表
    book_new = Workbook()
    # 删除默认的Sheet
    default_sheet = book_new["Sheet"]
    if default_sheet is not None:
        book_new.remove(default_sheet)
    # 使用文件名作为工作簿的名字
    compound_index = os.path.split(os.path.splitext(xls_path)[0])[-1]
    sheet_new = book_new.create_sheet(compound_index, 0)
    # 填入数据
    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet_new.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)
    xlsx_path = os.path.splitext(xls_path)[0] + ".xlsx"
    book_new.save(xlsx_path)
    return xlsx_path


def change_suffix(filepath: str, dst_suffix: str):
    """
    将filepath指定的文件后缀改为dst_suffix
    :param filepath:
    :param dst_suffix:
    :return: 修改后的文件路径字符串，若参数存在None，返回None
    """
    if dst_suffix is not None:
        dot = str.find(dst_suffix, ".")
        if dot == -1:
            dst_suffix = "." + dst_suffix
        if filepath is not None:
            return os.path.splitext(filepath)[0] + dst_suffix
    return None

class DataMerger:
    def __init__(self, constants_yml_filename):
        """
            初始化全局变量
        """
        self.__ymlfilename = constants_yml_filename
        # 保存化合物名与其mol文件路径以及图片路径的映射，优化制表的遍历效率
        self.__compound_name2mol_map = dict()
        self.__compound_name2img_map = dict()
        # 替换器官名映射表
        self.__deprecated_organ_names = ymlReader.get_deprecated_organ_names(self.__ymlfilename)
        # 无用器官名列表
        self.__denied_organ_names = ymlReader.get_denied_organ_names(self.__ymlfilename)
        # 整合的目标时间点
        self.__time_intervals = ymlReader.get_time_intervals(self.__ymlfilename)
        # 分散数据集中对于时间间隔一栏（首栏）的叫法不一，会影响数据的收集，用列表记录用于函数处理
        self.__denied_interval_markers = ymlReader.get_denied_intervals(self.__ymlfilename)
        # 整合的目标器官名列表
        self.__organ_lists = ymlReader.get_target_organ_names(self.__ymlfilename)
        # 数据集的mol文件路径
        self.__mol_files = []

        """
            目录初始化
        """
        # 原始数据存放目录
        cwd = os.getcwd()
        self.__raw_data_dir = os.path.join(cwd, 'data')
        if not os.path.exists(self.__raw_data_dir):
            os.makedirs(self.__raw_data_dir)
            raise FileExistsError(f"数据集目录未发现，已创建该目录：{self.__raw_data_dir}，请将数据集放入该目录后重新运行")

        # 数据集汇总表存放目录
        cur_time = datetime.now().strftime("%Y%m%d")
        # self.__result_dir = f"{cwd}\\result\\{cur_time}"
        self.__result_dir = os.path.join(f"{cwd}", "result", f"{cur_time}")
        if not os.path.exists(self.__result_dir):
            os.makedirs(self.__result_dir)

        # 设置日志文件输出
        log_file = os.path.join(self.__result_dir, "DataMerger_DEBUG.log")
        file_handler = logging.FileHandler(log_file, encoding='utf8')
        file_handler.setLevel('DEBUG')
        file_handler.setFormatter(formatter)
        log.addHandler(file_handler)

        # 数据集化合物图片存放目录
        self.saved_pic_dir = os.path.join(self.__result_dir, 'img')
        if not os.path.exists(self.saved_pic_dir):
            os.makedirs(self.saved_pic_dir)

        # 读取原始数据目录中的mol文件，存放化合物编号及对应文件名的映射
        data_list = os.listdir(self.__raw_data_dir)
        for file in data_list:
            if file.endswith(".mol"):
                mol_file = os.path.join(self.__raw_data_dir, file)
                compound_name = os.path.splitext(file)[0]
                self.__compound_name2mol_map[compound_name] = mol_file
                self.__mol_files.append(mol_file)

        # TODO: 修改输出文件名
        # 创建输出的excel文件
        self.output_excel_filepath = f"{self.__result_dir}\\数据表汇总.xlsx"
        if not os.path.exists(self.output_excel_filepath):
            wkc = Workbook(self.output_excel_filepath)
            wkc.save(self.output_excel_filepath)

        # 记录出错的文件
        self.errorfile = set()

    def start_merging(self):
        """
            启动数据整合
        """
        self.__get_imgs()
        main_df = self.__init_workbook_dataframe()
        # 遍历所有化合物对应的数据excel文件，整合到一个Dataframe中
        for compound_name, compound_file in tqdm(self.__compound_name2mol_map.items(), desc="正在遍历化合物数据"):
            if compound_name is not None:
                # 获得mol文件对应的excel文件并读取数据
                xlsx_filepath = change_suffix(compound_file, "xlsx")
                # 若xlsx文件不存在但存在xls文件，则转换xls为xlsx
                if xlsx_filepath is not None and not os.path.exists(xlsx_filepath):
                    xls_filepath = change_suffix(compound_file, "xls")
                    if not os.path.exists(xls_filepath):
                        log.error(f"化合物编号{compound_name}没有xls或xlsx数据表文件")
                        self.__save_error_compound(compound_name)
                        continue
                    log.info(f"化合物编号{compound_name}数据表格式为xls，另存为xlsx格式")
                    xlsx_filepath = xls2xlsx(xls_filepath)
                df = self.__get_DataFrame_from_workbook(xlsx_filepath)
                if df is not None:
                    try:
                        main_df = pd.concat([main_df, df], axis=0)
                    except pd.errors.InvalidIndexError as IIE:
                        log.debug(f"整合数据文件存在索引问题，对应化合物编号为{compound_name}")
                        log.debug(traceback.format_exc())
                        self.__save_error_compound(compound_file)
                        continue
                    except Exception as e:
                        log.error(f"整合数据文件出错，出错的化合物编号为{compound_name}")
                        log.error(traceback.format_exc())
                        self.__save_error_compound(compound_file)
        # 去重，并预留保存化合物结构图以及SMILES的空列后保存到excel文件中
        main_df = pd.DataFrame.dropna(main_df, axis=1, how='all')
        main_df.insert(loc=1, column='Compound structure', value="")
        main_df.insert(loc=1, column='SMILES', value="")
        main_df.to_excel(self.output_excel_filepath, index=False, engine='openpyxl', encoding='utf-8')
        log.info(f"完成化合物数据遍历，数据表保存至{self.output_excel_filepath}")

    def insert_SMILES_imgs(self):
        """
            使用openpyxl打开excel文件并进行设定
        """
        log.info("正在进行化合物结构图及SMILES插入工作，请勿打开数据表直到工作完成")
        # 打开数据汇总表
        wbc = openpyxl.load_workbook(self.output_excel_filepath)
        # 操作当前相应的表
        wsc = wbc.active

        # 调整第一二列的列宽，并调整第一行的行高
        wsc.column_dimensions['A'].width = 25
        wsc.column_dimensions['B'].width = 50
        wsc.row_dimensions[1].height = 30
        # 全表靠左纵向居中
        alignment = Alignment(horizontal='left', vertical='center')
        for col in wsc.columns:
            for cell in col:
                cell.alignment = alignment

        """
            插入SMILES
        """
        # 第一个化合物从工作簿的第2行开始，记录当前操作的行数
        row = 2
        SMILES_column = 2
        for compound_name_cell in tqdm(wsc['A'], desc="正在插入化合物SMILES: "):
            # 从第一列获取化合物名，从缓存的映射表中得到化合物对应的mol文件路径
            compound_file_name = self.__compound_name2mol_map.get(compound_name_cell.value)
            if compound_file_name is not None:
                try:
                    # 使用RDkit读取mol文件并计算SMILES
                    writer = Chem.MolFromMolFile(compound_file_name)
                    SMILES = Chem.MolToSmiles(writer)
                except OSError as ose:
                    log.debug(f"输入的mol文件存在问题，化合物编号为{compound_file_name}")
                    log.debug(traceback.format_exc())
                    self.__save_error_compound(compound_file_name)
                    row = row + 1
                    continue
                except Exception as e:
                    log.error(f"SMILES插入出错，化合物编号为{compound_file_name}")
                    log.error(traceback.format_exc())
                    self.__save_error_compound(compound_file_name)
                    row = row + 1
                    continue
                # 将SMILES填写到对应列中
                wsc.cell(row, SMILES_column).value = SMILES
                wsc.cell(row, SMILES_column).alignment = alignment
                row = row + 1

        """
            图片写入到汇总表
        """
        # 第一个化合物从工作簿的第2行开始，记录当前操作的行数
        row = 2
        # 对map长度的计数器，防止map内数据已经使用完的情况下程序还在对excel进行行遍历
        count = 0
        # map_length = len(self.compound_name2img_map)
        # 调整列宽
        wsc.column_dimensions['C'].width = 20

        # 读取A列的化合物名
        try:
            for compound_name_cell in tqdm(wsc['A'], desc="正在插入化合物结构图: "):
                # if count == map_length:
                #     break
                compound_name = compound_name_cell.value
                # 跳过第一行
                if compound_name == '文献编号' or compound_name == '化合物编号':
                    continue
                img_path = self.__compound_name2img_map.get(compound_name)
                if img_path is not None:
                    img = Image(img_path)
                    # img = PImage.open(img_path).resize((120, 120))

                    # 图片只保存在C列，只对C列每一行进行操作
                    wsc.add_image(img, 'C' + str(row))
                    # 调整行高
                    wsc.row_dimensions[row].height = 96
                    row = row + 1
                    count = count + 1
        except Exception as e:
            log.error(traceback.format_exc())
        finally:
            wbc.save(self.output_excel_filepath)
            log.info("插入工作完成，数据表保存成功")
        log.debug(f"存在问题的化合物编号: {self.errorfile}")

    def __get_imgs(self, size=(120, 120)):
        """
            把mol文件生成的化合物结构图保存到图片目录中，并保存对应化合物与图片文件路径的映射
        """
        # 读取数据集文件
        for mol_file in tqdm(self.__mol_files, desc="正在获取化合物结构图: "):
            # 将文件路径按照"路径+后缀名"进行拆分，确认文件后缀名是否为mol
            # 如"data\\44593-C6-7α-18FFES.mol"，拆分为('data\\44593-C6-7α-18FFES', '.mol')
            split_path = os.path.splitext(mol_file)
            if split_path[-1] == '.mol':
                # 将路径按照最后一个分隔符进行拆分，如'a\\data\\44593-C6-7α-18FFES'，拆分为('a\\data', '44593-C6-7α-18FFES')
                # 获得化合物名
                compound_name = os.path.split(split_path[0])[-1]
                try:
                    # 读取mol文件并生成化合物结构图
                    mol = Chem.MolFromMolFile(mol_file)
                    img_path = os.path.join(self.saved_pic_dir, compound_name + '.png')
                    # Draw.MolToImage(mol, size=(120, 120), kekulize=True)
                    Draw.MolToFile(mol, img_path, size=size)
                    # 保存对应化合物与图片文件路径的映射
                    self.__compound_name2img_map[compound_name] = img_path
                except OSError as ose:
                    log.debug(f"输入的mol文件存在问题，化合物编号为{compound_name}")
                    log.debug(traceback.format_exc())
                    self.__save_error_compound(compound_name)
                except Warning as e:
                    log.debug(f"生成化合物编号{compound_name}的结构图时产生警告: {traceback.format_exc()}")
                    self.__save_error_compound(compound_name)
                except Exception as e:
                    log.error(f"化合物编号{compound_name}的结构图生成出现问题:")
                    log.error(traceback.format_exc())
                    self.__save_error_compound(mol_file)
        log.info(f"化合物结构图处理完成，保存至目录: {self.saved_pic_dir}")

    def __init_workbook_dataframe(self):
        """
            初始化带全部列头的dataframe, 用于将读取到的数据对应列头并填入dataframe中

            Return:
                包含所有列头的空DataFrame
        """
        log.info("初始化Dataframe表")
        headers = ['Compound index']
        for organ in self.__organ_lists:
            for time in self.__time_intervals:
                headers.append(organ + " mean" + str(time) + "min")
                headers.append(organ + " sd" + str(time) + "min")
        df = pd.DataFrame(columns=headers)
        return df

    def __get_DataFrame_from_workbook(self, workbook):
        """
            用于从excel文件中读取单个药物的数据, 将数据进行处理并打包成DataFrame返回

            流程如下：

            1.记录列表头的时间点数据，进行预处理

            2.列表剩余数据对应药物在不同器官下的浓度数据，以{器官：浓度数据}的格式存储到字典中

            3.将所有的器官名与时间点数据组合为新的列表头

            4.创建新的Dataframe，使用3的列表头

            5.遍历2的字典中的浓度数据，根据器官与对应的时间结点，填入到Dataframe对应的位置上

            示例：

            化合物A.excel

            |       |30min  |60min  |

            |brain  |1      |2      |

            |blood  |0.1    |0.3    |

            ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓

            |compound_index |brain 30min|brain 60min|blood 30min|blood 60min|

            |化合物A         |1          |2          |0.1        |0.3        |

            Args:
                workbook: 总表对应excel文件的文件路径
            Return:
                根据时间点与器官名展开为一行的DataFrame，包含化合物名与浓度数据
        """
        # 读取已创建的excel工作簿
        try:
            wb = openpyxl.load_workbook(workbook)
        except FileNotFoundError as e:
            log.error(traceback.format_exc())
            return None
        worksheet = wb.active
        # 化合物编号
        compound_index = os.path.splitext(os.path.split(workbook)[-1])[0]
        # 保存器官与不同时间对应浓度的字典
        organ_concentration_dict = dict()
        # 判断处理的是否是第一行列表头，将浓度表格数据分为列表头和列表数据两部分分别处理
        is_header_row = True
        # 保存列表头数据（时间点数据）的列表
        time_headers = []

        for row in worksheet.rows:
            # 记录第一行的列头（保存着时间点数据），分析这些时间点数据，并将所有时间统一为分钟
            if is_header_row:
                # 遍历每列的时间点数据
                for cell in row:
                    if cell.value is None:
                        continue
                    time_header = str(cell.value).strip().replace(
                        " ", "").replace("\n", "").lower()
                    # 判断时间点是否是被拒绝的，是则跳过不处理，否则为正常的时间点数据
                    if time_header in self.__denied_interval_markers:
                        continue
                    # 修正由于OCR识别问题导致的字符错误，替换为正常的字符
                    error_text = ymlReader.get_OCR_error_text(self.__ymlfilename)
                    if error_text is not None and len(error_text) > 0:
                        # TODO: 有的时间点错误识别为"mea120min"，若在下面的替换规则中加入mea: mean，则会导致正常的"mean"被替换成"meann"
                        # TODO: 目前删除了mea: mean的替换规则，因此出现上述错误识别的数据无法修正，可以用新的数组记录出错的数据文件
                        for k, v in error_text.items():
                            time_header = time_header.replace(k, v)
                    # 存在部分时间点数据缺少时间单位，默认附上min
                    if not time_header.endswith("min") and not time_header.endswith("h"):
                        time_header = time_header + "min"
                    # 将单位是小时的时间点数据转换为分钟
                    if time_header[-1] == 'h':
                        try:
                            # 获取小时数字的字符串范围
                            index = time_header.find('mean')
                            if index != -1:
                                index = index + 4
                            else:
                                index = time_header.find('sd')
                                if index != -1:
                                    index = index + 2
                            # 转换为分钟
                            if index != -1:
                                hour = int(time_header[index:-1])
                                time_header = time_header[:index] + str(hour * 60) + 'min'
                            else:
                                log.error(
                                    f"时间点数据存在缺失，对应的化合物为{compound_index}，出错的时间点为{time_header}")
                                self.__save_error_compound(compound_index)
                                continue
                        except ValueError as e:
                            log.error(f"转换时间点数据出错，对应的化合物为{compound_index}，出错的时间点为{time_header}")
                            log.error(traceback.format_exc())
                            self.__save_error_compound(compound_index)
                    # 还存在部分时序列头的时间数字缺失，输出错误的数据并防止输入到总数据集中
                    if time_header != 'sdmin' and time_header != 'meanmin':
                        time_headers.append(time_header)
                    else:
                        log.error(f"时间点数据存在缺失，对应的化合物为{compound_index}，出错的时间点为{time_header}")
                        self.__save_error_compound(compound_index)
                # END: for cell in row:
                if len(time_headers) > 0:
                    # 部分数据文件中的数据并非从第一行开始，通过判断列表的长度可以充当跳过前面空行的作用
                    is_header_row = False
                    # 试图找出错误的时间列头的列表
                    if str(time_headers[0]).find('mean') == -1 and str(time_headers[0]).find('sd') == -1:
                        log.error(f"错误的列表头，对应的化合物为{compound_index}，列表头数据为{time_headers}")
                        self.__save_error_compound(compound_index)
            # END: if is_header_row:
            # 接着处理带数值的列表数据
            else:
                # 将一行的数据先置入列表中，再根据行名与行数据保存成字典形式
                temp_list = []
                for cell in row:
                    if cell.value is not None:
                        temp_list.append(str(cell.value).strip()
                                         .replace("_x0001_", "")
                                         .replace(" ", "")
                                         .replace("\n", ""))
                if len(temp_list) > 0:
                    # 每一行的第一列为器官名，提取出来作为字典的键，其他数据为值
                    organ_name = str(temp_list[0]).lower()
                    # 若器官名不合规或者是被取消的，跳过
                    if organ_name is not None and len(organ_name) > 0:
                        if organ_name.isalpha() and organ_name not in self.__denied_organ_names:
                            # 若器官名是需要被替换的，替换
                            if self.__deprecated_organ_names.get(organ_name) is not None:
                                organ_name = self.__deprecated_organ_names.get(organ_name)
                            organ_concentration_dict[organ_name] = temp_list[1:]
        # END: for row in worksheet.rows

        # 检查数据完整性
        if is_header_row is True or len(organ_concentration_dict) == 0:
            self.__save_error_compound(compound_index)
            raise ValueError(f"化合物 {compound_index} 数据存在问题")
        # 组合时间表头与器官名，用于置入DataFrame成为新的表头
        organs = list(organ_concentration_dict.keys())
        extended_headers = ['Compound index']
        for organ in organs:
            for time_header in time_headers:
                try:
                    # 组合新的表头，并添加到新表头列表中
                    extended_headers.append(str.lower(" ".join([str(organ), str(time_header)])))
                except Exception as e:
                    log.error(traceback.format_exc())
                    log.error(f"出错的化合物: {compound_index},器官名: {organ}, "
                              f"时间点: {time_header}, 当前替换后的列表头: {extended_headers}")
                    self.__save_error_compound(compound_index)
        # 设置DataFrame并写入化合物编号
        df = pd.DataFrame(columns=extended_headers)
        df[extended_headers[0]] = [compound_index]

        # 遍历器官数据，并写入到DataFrame对应的列中
        for organ_name, organ_data in organ_concentration_dict.items():
            cur = 0
            # print(f"器官{organ_name}：数据长度为{len(organ_data)}, 列头长度为{len(time_headers)}")
            for data in organ_data:
                try:
                    time_header = str.lower(' '.join([str(organ_name), str(time_headers[cur])]))
                    df[time_header] = [data]
                    cur = cur + 1
                # TODO: 重写错误信息输出
                except Exception as e:
                    self.__save_error_compound(compound_index)
                    log.error(f"Sheet rawdata: {organ_concentration_dict}")
                    log.error(f"Organs list: {organs}")
                    log.error(f"Headers list: {time_headers}")
                    log.error(f"Problem organ name: {organ_name}")
                    log.error(f"Problem organ rawdata: {data}")
                    log.error(f"Cursor index: {cur}")
                    log.error(f"Problem compound index: {compound_index}")
                    log.error(traceback.format_exc())
                    break
        return df

    def __save_error_compound(self, compound_index_or_filename: str):
        """
        记录以及格式化出错的化合物编号到errorfile中
        :param compound_index_or_filename:
        :return:
        """
        if compound_index_or_filename is None:
            return
        split_path = os.path.splitext(compound_index_or_filename)
        if split_path[-1] == '.mol':
            # 获得化合物名
            compound_name = os.path.split(split_path[0])[-1]
            self.errorfile.add(compound_name)
        else:
            self.errorfile.add(compound_index_or_filename)